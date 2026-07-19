import numpy as np
import os
import sys
import zipfile
import csv
from PySide6.QtWidgets import QDialog, QFileDialog, QMessageBox, QInputDialog, QTableWidgetItem

from information import COMBINED_INPUTS



def _get_base_dir():
    """Get the base directory for data files (handles both script and exe contexts)."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def is_valid_number(value):
    try:
        return float(value) >= 0
    except (ValueError, TypeError):
        return False
    

def _normalize_column_name(name):
    return str(name).strip().lower().replace(" ", "_").replace("-", "_")


def _get_target_spreadsheet(state, sheet_name=None):
    """Return the spreadsheet widget for the current app state."""
    if state is None:
        return None

    if hasattr(state, "table") and hasattr(state, "headers") and hasattr(state, "scenario_data"):
        return state

    if hasattr(state, "current_spreadsheet"):
        spreadsheet = state.current_spreadsheet()
        if spreadsheet is not None:
            return spreadsheet

    if hasattr(state, "page"):
        page = None
        if isinstance(getattr(state, "page", None), dict):
            page = state.page.get("LIBPage")
            if page is None:
                for candidate in state.page.values():
                    if hasattr(candidate, "tabs") and hasattr(candidate, "add_sheet"):
                        page = candidate
                        break
        elif hasattr(state, "tabs") and hasattr(state, "add_sheet"):
            page = state

        if page is not None:
            if sheet_name is not None and hasattr(page, "add_sheet"):
                page.add_sheet()
                spreadsheet = page.current_spreadsheet()
                if spreadsheet is not None and hasattr(page, "tabs"):
                    index = page.tabs.indexOf(spreadsheet)
                    if index >= 0:
                        page.tabs.setTabText(index, sheet_name)
                return spreadsheet

            if hasattr(page, "current_spreadsheet"):
                spreadsheet = page.current_spreadsheet()
                if spreadsheet is None and hasattr(page, "add_sheet"):
                    page.add_sheet()
                    spreadsheet = page.current_spreadsheet()
                if spreadsheet is not None:
                    return spreadsheet

    if hasattr(state, "tabs"):
        widget = state.tabs.currentWidget()
        if hasattr(widget, "table") and hasattr(widget, "headers"):
            return widget

    return None


def _ensure_spreadsheet_capacity(spreadsheet, required_rows):
    from calculations import is_string_field

    if required_rows <= spreadsheet._full_scenario_data.shape[0]:
        return

    new_size = max(required_rows, spreadsheet._full_scenario_data.shape[0] + 10)
    new_data = np.zeros(new_size, dtype=spreadsheet.dtype)
    for header in spreadsheet.headers:
        if not is_string_field(header):
            new_data[header] = np.nan

    old_data = spreadsheet._full_scenario_data
    for header in spreadsheet.headers:
        new_data[header][:old_data.shape[0]] = old_data[header]

    spreadsheet._full_scenario_data = new_data
    spreadsheet.table.setRowCount(new_size)


def _populate_spreadsheet_from_dataframe(spreadsheet, df):
    import pandas as pd
    from calculations import is_string_field

    headers = list(spreadsheet.headers)
    header_indices = {header: idx for idx, header in enumerate(headers)}

    spreadsheet.table.blockSignals(True)
    try:
        for _, row in df.iterrows():
            target_row = None
            for candidate_row in range(spreadsheet._full_scenario_data.shape[0]):
                if not spreadsheet._row_has_data(candidate_row):
                    target_row = candidate_row
                    break

            if target_row is None:
                _ensure_spreadsheet_capacity(spreadsheet, spreadsheet._full_scenario_data.shape[0] + 1)
                target_row = spreadsheet._full_scenario_data.shape[0] - 1

            for header in headers:
                col_name = None
                if header in df.columns:
                    col_name = header
                else:
                    normalized = _normalize_column_name(header)
                    matches = [column for column in df.columns if _normalize_column_name(column) == normalized]
                    if matches:
                        col_name = matches[0]

                value = row.get(col_name, "") if col_name is not None else ""
                col_index = header_indices[header]

                if is_string_field(header):
                    if pd.isna(value) or value is None:
                        text = ""
                    else:
                        text = str(value)
                    spreadsheet._full_scenario_data[header][target_row] = text
                    item = spreadsheet.table.item(target_row, col_index)
                    if item is None:
                        item = QTableWidgetItem()
                    item.setText(text)
                    spreadsheet.table.setItem(target_row, col_index, item)
                else:
                    if value in ("", None) or (isinstance(value, float) and np.isnan(value)) or pd.isna(value):
                        numeric_value = np.nan
                    else:
                        try:
                            numeric_value = float(value)
                        except (TypeError, ValueError):
                            numeric_value = np.nan

                    spreadsheet._full_scenario_data[header][target_row] = numeric_value
                    item = spreadsheet.table.item(target_row, col_index)
                    if item is None:
                        item = QTableWidgetItem()
                    item.setText("" if np.isnan(numeric_value) else str(numeric_value))
                    spreadsheet.table.setItem(target_row, col_index, item)

            spreadsheet._refresh_scenario_data_view()
    finally:
        spreadsheet.table.blockSignals(False)


def _load_single_sheet(state, df, sheet_name=None):
    """Load a single DataFrame into the current spreadsheet widget/tab."""
    spreadsheet = _get_target_spreadsheet(state, sheet_name=sheet_name)
    if spreadsheet is None:
        QMessageBox.critical(None, "Import Error", "No active spreadsheet is available for import.")
        return getattr(state, "scenario_counter", 0), []

    _populate_spreadsheet_from_dataframe(spreadsheet, df)
    return len(spreadsheet.scenario_data), []


def load_file(state, file_path=None):
    import pandas as pd

    if file_path is None or not os.path.exists(file_path):
        file_path, _ = QFileDialog.getOpenFileName(
            filter="Excel and CSV files (*.csv *.xlsx *.xls)",
            caption="Select a CSV or Excel File"
        )

    if not file_path:
        print("DEBUG: No file selected, returning")
        return getattr(state, "scenario_counter", 0)

    try:
        lower_path = str(file_path).lower()
        if lower_path.endswith(".csv"):
            df = pd.read_csv(file_path)
            if df.empty:
                QMessageBox.warning(None, "Empty File", "The selected file is empty.")
                return getattr(state, "scenario_counter", 0)

            counter, skipped = _load_single_sheet(state, df)
            if hasattr(state, 'update_status'):
                state.update_status(f"Loaded {len(df)} scenarios from file", len(df))

            loaded_count = len(df) - len(skipped)
            if skipped:
                QMessageBox.information(None, "Load Complete",
                    f"Successfully loaded {loaded_count} scenarios\n"
                    f"Skipped {len(skipped)} rows with invalid data (rows: {', '.join(map(str, skipped))})")
            else:
                QMessageBox.information(None, "Load Complete", f"Successfully loaded {loaded_count} scenarios")
            return counter
        elif lower_path.endswith(".zip"):
            with zipfile.ZipFile(file_path, 'r') as archive:
                csv_files = [name for name in archive.namelist() if name.lower().endswith('.csv')]
                if not csv_files:
                    QMessageBox.warning(None, "No CSV Found", "No CSV files were found in the selected archive.")
                    return getattr(state, "scenario_counter", 0)
                with archive.open(csv_files[0]) as handle:
                    df = pd.read_csv(handle)
            if df.empty:
                QMessageBox.warning(None, "Empty File", "The selected archive contains no data.")
                return getattr(state, "scenario_counter", 0)
            counter, skipped = _load_single_sheet(state, df)
            if hasattr(state, 'update_status'):
                state.update_status(f"Loaded {len(df)} scenarios from archive", len(df))
            loaded_count = len(df) - len(skipped)
            if skipped:
                QMessageBox.information(None, "Load Complete",
                    f"Successfully loaded {loaded_count} scenarios\n"
                    f"Skipped {len(skipped)} rows with invalid data (rows: {', '.join(map(str, skipped))})")
            else:
                QMessageBox.information(None, "Load Complete", f"Successfully loaded {loaded_count} scenarios")
            return counter
        else:
            xl = pd.ExcelFile(file_path)
            sheet_names = xl.sheet_names

            if not sheet_names:
                QMessageBox.warning(None, "Empty File", "The selected Excel file has no sheets.")
                return getattr(state, "scenario_counter", 0)

            total_loaded = 0
            total_skipped = []

            for sheet_name in sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                if df.empty:
                    continue
                _, skipped = _load_single_sheet(state, df, sheet_name=sheet_name)
                total_loaded += len(df) - len(skipped)
                total_skipped.extend(skipped)

            if hasattr(state, 'update_status'):
                state.update_status(f"Loaded {total_loaded} scenarios from {len(sheet_names)} sheet(s)", total_loaded)

            if total_skipped:
                QMessageBox.information(None, "Load Complete",
                    f"Successfully loaded {total_loaded} scenarios across {len(sheet_names)} sheet(s)\n"
                    f"Skipped {len(total_skipped)} rows with invalid data")
            else:
                QMessageBox.information(None, "Load Complete", f"Successfully loaded {total_loaded} scenarios across {len(sheet_names)} sheet(s)")
            return total_loaded

    except Exception as e:
        QMessageBox.critical(None, "Load Error", f"An error occurred while loading the file:\n{str(e)}")
        return getattr(state, "scenario_counter", 0)
    
    
def generate_input_template(state):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    from PySide6.QtWidgets import QFileDialog, QMessageBox

    try:
        spreadsheet = _get_target_spreadsheet(state)
        if spreadsheet is not None and hasattr(spreadsheet, "headers"):
            columns = list(spreadsheet.headers)
        else:
            columns = list(COMBINED_INPUTS)

        if "Scenario" not in columns:
            columns = ["Scenario"] + columns

        df = pd.DataFrame(columns=columns)

        parent = state if hasattr(state, "parent") else None
        file_path, _ = QFileDialog.getSaveFileName(
            parent,
            "Save Template As",
            "bat_offgassing_input_template.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        if not str(file_path).lower().endswith(".xlsx"):
            file_path = f"{file_path}.xlsx"

        df.to_excel(file_path, index=False)

        wb = load_workbook(file_path)
        ws = wb.active

        for cell in ws[1]:
            cell.alignment = Alignment(wrap_text=True)

        ws.row_dimensions[1].height = 30

        for col_idx, col in enumerate(ws.columns, start=1):
            max_length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(max_length + 2, 10)

        wb.save(file_path)

        QMessageBox.information(None, "Template Created", f"Excel input template saved to:\n{file_path}")

    except Exception as e:
        QMessageBox.critical(None, "Template Error", f"An error occurred while creating the template:\n{str(e)}")
        

def import_gas_flowrate_data(state):
    """Import gas flowrate data from an Excel file with four tabs (co, h2, total_hydrocarbons, co2).
    Each tab should contain two columns: Column A = time (s), Column B = flowrate (m³/s).
    Time data can be at any interval and does not need to start at 0.
    The data is interpolated onto a uniform 1-second grid starting at 0s.
    Stores the resulting arrays on state for use by flammability_assessment_calc.
    """
    import pandas as pd
    import numpy as np

    file_path, _ = QFileDialog.getOpenFileName(
        state if hasattr(state, "parent") else None,
        "Select Gas Flowrate Data Excel File",
        "",
        "Excel files (*.xlsx *.xls)"
    )
    if not file_path:
        return  # User cancelled

    # Ask user for the off-gassing start time (in minutes) within the data
    offgas_start_min = QInputDialog.getDouble(
        "Off-gassing Start Time",
        "Enter the time (in minutes) at which off-gassing begins in the data:\n"
        "(Data before this time will be discarded)",
        initialvalue=0.0,
        minvalue=0.0
    )
    if offgas_start_min is None:
        return  # User cancelled

    expected_tabs = ["co", "h2", "total_hydrocarbons", "co2"]

    try:
        xl = pd.ExcelFile(file_path)
        sheet_names_lower = {name.strip().lower(): name for name in xl.sheet_names}

        # Validate that all expected tabs exist
        missing_tabs = [tab for tab in expected_tabs if tab not in sheet_names_lower]
        if missing_tabs:
            QMessageBox.critical(None,
                "Missing Tabs",
                f"The selected Excel file is missing the following required tabs:\n{', '.join(missing_tabs)}\n\n"
                f"Found tabs: {', '.join(xl.sheet_names)}"
            )
            return

        flowrate_data = {}
        max_time = 0  # Track the longest duration across all gases

        for tab in expected_tabs:
            actual_sheet_name = sheet_names_lower[tab]
            df = xl.parse(actual_sheet_name)
            if df.empty or df.shape[1] < 2:
                QMessageBox.critical(None,
                    "Invalid Tab Format",
                    f"The tab '{actual_sheet_name}' must have at least two columns:\n"
                    f"Column A = Time (s), Column B = Flowrate (m³/s)"
                )
                return

            # Read first two columns: time and flowrate
            time_raw = pd.to_numeric(df.iloc[:, 0], errors='coerce')
            flow_raw = pd.to_numeric(df.iloc[:, 1], errors='coerce')

            # Drop rows where either value is NaN
            valid_mask = time_raw.notna() & flow_raw.notna()
            time_vals = time_raw[valid_mask].values
            flow_vals = flow_raw[valid_mask].values

            # Convert flowrate from L/min to m³/s: divide by 60 (L/s) then by 1000 (m³/s)
            flow_vals = flow_vals / 60.0 / 1000.0

            if len(time_vals) < 2:
                QMessageBox.critical(None,
                    "Insufficient Data",
                    f"The tab '{actual_sheet_name}' has fewer than 2 valid data points after cleaning."
                )
                return

            # Sort by time in case data is not ordered
            sort_idx = np.argsort(time_vals)
            time_vals = time_vals[sort_idx]
            flow_vals = flow_vals[sort_idx]

            # Track the maximum end time across all gases (in minutes)
            if time_vals[-1] > max_time:
                max_time = time_vals[-1]

            flowrate_data[tab] = (time_vals, flow_vals)

        # Create uniform 1-unit grid in minutes, then convert to seconds
        # Interpolate at 1-minute intervals first, then expand to 1-second grid
        max_time_minutes = max_time
        # Create a grid at the original minute scale with 1/60 minute steps (= 1 second)
        max_time_seconds = int(np.ceil(max_time_minutes * 60))
        uniform_time = np.arange(0, max_time_seconds + 1, 1)  # 0, 1, 2, ... seconds
        # Corresponding time in minutes for interpolation
        uniform_time_minutes = uniform_time / 60.0

        # Interpolate each gas onto the uniform grid (interpolate in minutes, result is per-second)
        interpolated_data = {}
        for tab in expected_tabs:
            time_vals, flow_vals = flowrate_data[tab]

            # Linear interpolation using minute-scale time points
            interpolated = np.interp(uniform_time_minutes, time_vals, flow_vals, left=0.0, right=0.0)
            interpolated = np.maximum(interpolated, 0.0)  # Clamp any negative values to zero
            interpolated_data[tab] = interpolated
            print(f"  {tab}: {len(time_vals)} raw points -> {len(interpolated)} points "
                  f"(time range {time_vals[0]:.2f}min - {time_vals[-1]:.2f}min = {time_vals[-1]*60:.0f}s)")

        # Trim data to start at the user-specified off-gassing start time
        offgas_start_seconds = int(round(offgas_start_min * 60))
        if offgas_start_seconds > 0:
            for tab in expected_tabs:
                if offgas_start_seconds < len(interpolated_data[tab]):
                    interpolated_data[tab] = interpolated_data[tab][offgas_start_seconds:]
                else:
                    interpolated_data[tab] = np.array([0.0])
            uniform_time = np.arange(0, len(interpolated_data[expected_tabs[0]]))
            print(f"  Off-gassing start offset applied: trimmed first {offgas_start_seconds}s ({offgas_start_min:.2f} min)")
        else:
            print("  No off-gassing start offset (begins at time 0)")

        # Store on state
        state.gas_flowrate_data = interpolated_data
        count = len(uniform_time)

        # Show data review popup (Qt app supplies its own via state hook)
        review = getattr(state, "flowrate_review_popup", None) or _show_flowrate_review_popup
        review(interpolated_data, uniform_time, expected_tabs)

        max_time_seconds = len(uniform_time) - 1
        print(f"Gas flowrate data loaded: {count} time steps (0-{max_time_seconds}s) from {file_path}")

    except Exception as e:
        QMessageBox.critical(None, "Import Error", f"Failed to import gas flowrate data:\n{str(e)}")
        
def _show_flowrate_review_popup(interpolated_data, uniform_time, gas_labels):
    """Display a popup showing the interpolated flowrate data for each gas for user review."""
    import PySide6.QtWidgets as QtWidgets

    popup = QtWidgets.QDialog()
    popup.setWindowTitle("Imported Flowrate Data Review")
    popup.resize(900, 500)

    # Header
    header_label = QtWidgets.QLabel("Interpolated Gas Flowrate Data (1-second intervals)", popup)
    header_label.setFont(QtWidgets.QFont('Segoe UI', 11, QtWidgets.QFont.Bold))
    header_label.setAlignment(QtWidgets.Qt.AlignCenter)
    header_label.setContentsMargins(0, 10, 0, 5)
    header_label.show()

    subheader_label = QtWidgets.QLabel(f"Total duration: {len(uniform_time)} seconds  |  Gases: {', '.join(gas_labels)}", popup)
    subheader_label.setFont(QtWidgets.QFont('Segoe UI', 9))
    subheader_label.setAlignment(QtWidgets.Qt.AlignCenter)
    subheader_label.setContentsMargins(0, 0, 0, 10)
    subheader_label.show()

    # Notebook with a tab per gas
    notebook = QtWidgets.QTabWidget(popup)
    notebook.setGeometry(10, 50, 880, 400)

    for gas in gas_labels:
        frame = QtWidgets.QWidget()
        notebook.addTab(frame, gas.upper().replace("_", " "))

        # Treeview for data display
        tree_frame = QtWidgets.QWidget(frame)
        tree_layout = QtWidgets.QVBoxLayout(tree_frame)
        tree_frame.setLayout(tree_layout)

        columns = ("time", "flowrate")
        tree = QtWidgets.QTreeWidget(tree_frame)
        tree.setColumnCount(2)
        tree.setHeaderLabels(["Time (s)", "Flowrate (m³/s)"])
        tree.setColumnWidth(0, 120)
        tree.setColumnWidth(1, 200)

        tree_layout.addWidget(tree)

        # Populate with data
        data = interpolated_data[gas]
        for i, t in enumerate(uniform_time):
            tree.insert("", "end", values=(int(t), f"{data[i]:.8e}"))

        # Summary label at bottom of each tab
        non_zero = data[data > 0]
        summary = (f"Points: {len(data)}  |  "
                   f"Max: {data.max():.6e} m³/s  |  "
                   f"Non-zero: {len(non_zero)}  |  "
                   f"Total volume: {data.sum():.6e} m³")
        summary_label = QtWidgets.QLabel(summary, frame)
        summary_label.setFont(QtWidgets.QFont('Segoe UI', 8))
        summary_label.setAlignment(QtWidgets.Qt.AlignCenter)
        summary_label.setContentsMargins(0, 5, 0, 5)
        summary_label.show()
    # Close button
    close_button = QtWidgets.QPushButton("Close", popup)
    close_button.clicked.connect(popup.destroy)
    close_button.setStyleSheet("background-color: #3498db; color: white; font: 9pt 'Segoe UI'; font-weight: bold;")
    close_button.setFixedHeight(30)
    close_button.move(400, 460)
    popup.exec()